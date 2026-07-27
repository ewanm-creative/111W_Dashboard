"""
111W_update_dashboard.py
Reads 111W_Logistics Output.xlsx and injects a DATA block into
111W_Container_Dashboard.html between the markers:
  // DATA_START
  // DATA_END

Run with --force to publish even if the sanity guard trips.
"""

import re, json, sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

HERE    = Path(__file__).parent
HTML_IN = HERE / "111W_Container_Dashboard.html"
HTML_OUT = HTML_IN
LOG_OUT = HERE / "111W_update_log.txt"

args = [a for a in sys.argv[1:] if not a.startswith("--")]
FORCE = "--force" in sys.argv

if args:
    XL = Path(args[0])
else:
    XL = HERE / "111W_Logistics Output.xlsx"

if not XL.exists():
    sys.exit("ERROR: Excel file not found: " + str(XL))
if not HTML_IN.exists():
    sys.exit("ERROR: HTML file not found: " + str(HTML_IN))

WARNINGS = []
def warn(msg):
    WARNINGS.append(msg)
    print("  ! " + msg)

print("Reading " + XL.name + " ...")
wb = load_workbook(XL, data_only=True)
print("  Sheets: " + str(wb.sheetnames))

def _d(dt): return str(dt.day)

def clean(v):
    if v is None: return "---"
    if isinstance(v, datetime): return str(v.month) + "/" + _d(v) + "/" + v.strftime("%Y")
    return str(v).strip() or "---"

def as_date(v):
    """Return a datetime if the cell holds a real date, else None."""
    if isinstance(v, datetime): return v
    if v is None: return None
    s = str(v).strip()
    if s in ("", "None", "---", "--"): return None
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try: return datetime.strptime(s, f)
        except ValueError: pass
    return None

def fmt_date(v):
    d = as_date(v)
    if d: return d.strftime("%b ") + _d(d) + d.strftime(", %Y")
    return clean(v)

def iso_date(v):
    d = as_date(v)
    return d.strftime("%Y-%m-%d") if d else None

# ── Sheet selection ────────────────────────────────────────────────
# Prefer the most specific name match so a stray sheet called e.g.
# "Container Notes" cannot hijack the container schedule.
def pick_sheet(*ranked_terms):
    for terms in ranked_terms:
        for name in wb.sheetnames:
            nl = name.strip().lower()
            if all(t in nl for t in terms):
                return name
    return None

SH_CTN = pick_sheet(
    ("ctnr", "schedule"), ("container", "schedule"), ("logistics", "schedule"),
    ("ctn", "schedule"), ("schedule",), ("ctnr",), ("container",),
)
if not SH_CTN:
    SH_CTN = wb.sheetnames[0]
    warn("No container schedule sheet found by name; falling back to '" + SH_CTN + "'")

print("  Container sheet: " + SH_CTN)
ws = wb[SH_CTN]

# ── Header detection ───────────────────────────────────────────────
NUM_HDRS   = ("CONT. #", "CTN #", "CONTAINER #", "CTN#", "CTNR #", "CTNR NO.", "CONT #")
STAT_HDRS  = ("CONT. STATUS", "STATUS", "CTNR STATUS")

header_row = None
headers = {}
for row in ws.iter_rows(min_row=1, max_row=20):
    vals = [str(c.value).strip().upper() for c in row if c.value]
    has_num    = any(v in NUM_HDRS for v in vals)
    has_status = any(v in STAT_HDRS for v in vals)
    has_units  = any(v == "UNITS" for v in vals)
    if has_num and (has_status or has_units):
        header_row = row[0].row
        break
if not header_row:
    header_row = 1
    warn("Could not find a header row; assuming row 1")

for cell in ws[header_row]:
    if cell.value:
        headers[str(cell.value).strip().upper()] = cell.column - 1

print("  Headers (row " + str(header_row) + "): " + str(list(headers.keys())[:10]))

def col(label, *names):
    for n in names:
        if n.upper() in headers:
            return headers[n.upper()]
    warn("Column not found for " + label + " (looked for: " + ", ".join(names) + ") - that field will be blank")
    return None

i_num   = col("container #", *NUM_HDRS)
i_st    = col("status", *STAT_HDRS)
i_week  = col("week", "PICKUP WEEK", "LOAD WK", "LOAD WEEK", "WK", "WEEK")
i_load  = col("load date", "LOAD DATE", "FACTORY LOAD DATE", "LOAD")
i_ship  = col("ship date", "SHIP DATE", "ETD", "DEPARTURE DATE")
i_vsl   = col("vessel", "VESSEL", "VESSEL NAME")
i_port  = col("port arrival", "ARRIVAL DATE", "PORT ARRIVAL", "ETA PORT", "ETA NY")
i_del   = col("delivery date", "DELIVERY DATE", "DELIVERY", "EST. DELIVERY", "DELIVERY ETA")
i_conf  = col("confirmed", "CONFIRM", "CONFIRMED?", "CONFIRMED")
i_qty   = col("unit qty", "UNIT Q.TY", "UNIT QTY", "QTY", "UNIT QUANTITY")
i_units = col("units", "UNITS", "UNIT RANGE", "UNIT NUMBERS", "UNIT LIST")
i_floors = col("floors", "FLOORS", "FLOOR", "FLOOR #")
i_notes = None
for n in ("MISC NOTES", "NOTES", "NOTE", "COMMENTS"):
    if n in headers:
        i_notes = headers[n]; break

# ── Unit range parsing ─────────────────────────────────────────────
ERR_TOKENS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")

def expand_ranges(s, ctx=""):
    """Expand '301:313,315:321' into unit ids.
    Accepts , and ; as separators, : and - as range markers, tolerates
    spaces and reversed ranges, and warns on anything it cannot read."""
    ids = []
    if s is None: return ids
    s = str(s).strip()
    if s in ("", "---", "--", "None"): return ids
    for part in re.split(r"[,;]", s):
        t = part.strip()
        if not t: continue
        # (?![A-Za-z]) so a label like "3rd Floor Doors" is not read as unit 3
        m = re.match(r"^(\d+)\s*[:\-]\s*(\d+)(?![A-Za-z0-9])\s*(.*)$", t)
        if m:
            a, b, tail = int(m.group(1)), int(m.group(2)), m.group(3).strip()
            if a > b: a, b = b, a
            if b - a > 500:
                warn(ctx + "range '" + t + "' spans " + str(b - a + 1) + " units - ignored as a likely typo")
                continue
            ids.extend(range(a, b + 1))
            if tail: warn(ctx + "ignored trailing text in '" + t + "'")
            continue
        m = re.match(r"^(\d+)(?![A-Za-z0-9])\s*(.*)$", t)
        if m:
            ids.append(int(m.group(1)))
            if m.group(2).strip(): warn(ctx + "ignored trailing text in '" + t + "'")
            continue
        warn(ctx + "could not read unit token '" + t + "' - those units are NOT tracked")
    return ids

# ── Containers ─────────────────────────────────────────────────────
STATUS_MAP = {"PROJ", "LDG", "LDD", "ENR", "INPRT", "D", "HOLD"}
STATUS_NORM = {
    "PROJ":"PROJ","LDG":"LDG","LDD":"LDD","ENR":"ENR","INPRT":"INPRT","D":"D","HOLD":"HOLD",
    "LOADING":"LDG","LOADED":"LDD",
    "EN ROUTE":"ENR","ENROUTE":"ENR","SEA FREIGHT":"ENR","SEA":"ENR",
    "PORT":"INPRT","IN PORT":"INPRT","AT PORT":"INPRT","PLD":"INPRT","PULLED":"INPRT",
    "DELIVERED":"D","DELIVERY":"D",
    "ON HOLD":"HOLD","CANCELLED":"HOLD","CANCELED":"HOLD","CUSTOMS":"HOLD",
    "DELAYED":"HOLD","STOPPED":"HOLD","TBD":"HOLD",
}

CONTAINERS = []
seen_nums = {}

for row in ws.iter_rows(min_row=header_row+1, values_only=True):
    if not any(row): continue
    def g(i, r=row): return r[i] if i is not None and i < len(r) else None
    num_raw = g(i_num)
    if num_raw is None: continue
    num_s = str(num_raw).strip()
    if num_s in ("", "None"): continue
    if num_s.lower() in ("example", "ex"): continue
    if num_s.upper() in ERR_TOKENS:
        warn("Row skipped: container # is '" + num_s + "'")
        continue

    # Keep non-numeric IDs (Doors, H1, 3A ...) as-is; sort numerics first.
    try:
        f = float(num_s)
        if f == 0: continue                      # spacer / all-zero rows
        num_disp = str(int(f)) if f == int(f) else num_s
        seq = (0, f, 0)
        is_num = True
    except ValueError:
        num_disp = num_s
        m = re.match(r"^(\d+)", num_s)
        seq = (1, float(m.group(1)) if m else 0.0, len(CONTAINERS))
        is_num = False

    if num_disp in seen_nums:
        warn("Duplicate container # '" + num_disp + "' appears more than once")
    seen_nums[num_disp] = True

    st_raw = clean(g(i_st)).upper().strip()
    st_norm = STATUS_NORM.get(st_raw, st_raw)
    if st_norm not in STATUS_MAP:
        if st_raw not in ("---", ""):
            warn("Container " + num_disp + ": unrecognised status '" + st_raw + "' - shown as On Hold")
            st_norm = "HOLD"
        else:
            st_norm = "PROJ"

    ctx = "Container " + num_disp + ": "
    units_raw = clean(g(i_units))
    unit_ids = expand_ranges(g(i_units), ctx)

    qty_raw = g(i_qty)
    try:
        qty = int(float(str(qty_raw)))
    except (TypeError, ValueError):
        qty = len(unit_ids)
        if qty_raw not in (None, "", "---"):
            warn(ctx + "unit qty '" + str(qty_raw) + "' is not a number - using the " + str(qty) + " units listed")
    if unit_ids and qty != len(unit_ids):
        warn(ctx + "unit qty says " + str(qty) + " but " + str(len(unit_ids)) + " units are listed")

    conf_raw = g(i_conf)
    confirmed = bool(conf_raw and str(conf_raw).strip().lower() in ("yes","y","confirmed","true","1","x"))

    note = ""
    if i_notes is not None:
        nv = g(i_notes)
        nvs = "" if nv is None else str(nv).strip()
        # Only treat it as a note if it reads like prose - that column also
        # carries bare numbers that are not comments.
        if nvs not in ("", "---", "--", "None") and re.search(r"[A-Za-z]", nvs):
            note = nvs

    CONTAINERS.append({
        "num": num_disp, "seq": seq, "isNum": is_num,
        "status": st_norm, "week": clean(g(i_week)),
        "loadDate": fmt_date(g(i_load)), "shipDate": fmt_date(g(i_ship)),
        "vessel": clean(g(i_vsl)), "portArrival": fmt_date(g(i_port)),
        "delivery": fmt_date(g(i_del)), "deliveryISO": iso_date(g(i_del)),
        "confirmed": confirmed, "unitQty": qty, "units": units_raw,
        "unitIds": unit_ids, "floors": clean(g(i_floors)), "note": note,
        "code": "—", "delivTime": "—",
    })

CONTAINERS.sort(key=lambda c: c["seq"])
print("  -> " + str(len(CONTAINERS)) + " containers parsed (" +
      str(sum(1 for c in CONTAINERS if not c["isNum"])) + " non-numeric IDs kept)")

for c in CONTAINERS:
    if c["deliveryISO"] is None and c["delivery"] != "---":
        warn("Container " + c["num"] + ": delivery '" + c["delivery"] +
             "' is not a real date - it will not appear in Next Up")

owner = {}
for c in CONTAINERS:
    for uid in c["unitIds"]:
        if uid in owner:
            warn("Unit " + str(uid) + " is listed in both container " + owner[uid] + " and " + c["num"])
        owner[uid] = c["num"]

# ── Unit-level overrides ───────────────────────────────────────────
UNIT_STATUS = {}
for sname in wb.sheetnames:
    if "unit status" in sname.lower() or "exceptions" in sname.lower():
        wsu = wb[sname]
        for row in wsu.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1 and row[0] is not None and row[1] is not None:
                try:
                    uid = int(float(str(row[0]).strip()))
                except (TypeError, ValueError):
                    continue
                st = STATUS_NORM.get(str(row[1]).strip().upper(), str(row[1]).strip().upper())
                if st in STATUS_MAP:
                    UNIT_STATUS[uid] = st
                else:
                    warn("Unit override " + str(uid) + ": unrecognised status '" + str(row[1]).strip() + "' - ignored")
        print("  -> Unit overrides: " + str(len(UNIT_STATUS)))
        break

# ── Transit performance ────────────────────────────────────────────
TRANSIT_PERF = []
SH_DASH = pick_sheet(("dashboard",), ("transit",))
if SH_DASH:
    wsd = wb[SH_DASH]
    for row in wsd.iter_rows(min_row=12, max_row=16, values_only=True):
        if not row: continue
        label = clean(row[0]) if len(row) > 0 else "---"
        value = clean(row[1]) if len(row) > 1 else "---"
        if label in ("---", "", "None") and len(row) > 1:
            label = clean(row[1])
            value = clean(row[3]) if len(row) > 3 else "---"
        if label not in ("---", "", "None"):
            TRANSIT_PERF.append({"label": label, "value": value})
    # A label that is just a number means we read the wrong sheet.
    if TRANSIT_PERF and all(str(r["label"]).replace(".", "").isdigit() for r in TRANSIT_PERF):
        warn("Transit performance rows look wrong (read from '" + SH_DASH + "') - dropped")
        TRANSIT_PERF = []
    print("  -> Transit perf: " + str(len(TRANSIT_PERF)) + " rows")

# ── Building matrix ────────────────────────────────────────────────
BUILDING_MATRIX = []
SH_BM = None
for sname in wb.sheetnames:
    nl = sname.strip().lower()
    if "building matrix" in nl and "(2)" not in nl and "condensed" not in nl:
        SH_BM = sname
        break

if SH_BM:
    wsbm = wb[SH_BM]
    print("  Scanning building matrix: " + SH_BM)

    live_status = {}
    for c in CONTAINERS:
        for uid in c["unitIds"]:
            live_status[uid] = c["status"]
    live_status.update(UNIT_STATUS)

    def as_floor(v):
        if v is None: return None
        try:
            f = int(float(str(v).strip()))
        except (TypeError, ValueError):
            return None
        return f if 1 <= f <= 99 else None

    # Locate floor blocks by scanning column B for a floor number that has
    # matching unit ids on the row above, instead of assuming row 7 + step 3.
    max_col = min(wsbm.max_column, 200)
    seen_floors = {}
    for r in range(1, wsbm.max_row + 1):
        floor_num = as_floor(wsbm.cell(row=r, column=2).value)
        if floor_num is None: continue
        row_ids, row_ktype, row_status = r - 1, r, r + 1
        if row_ids < 1 or row_status > wsbm.max_row: continue

        units = []
        for c_idx in range(3, max_col + 1):
            uid_val = wsbm.cell(row=row_ids, column=c_idx).value
            if uid_val is None: continue
            uid_str = str(uid_val).strip()
            if not uid_str.isdigit(): continue
            uid_int = int(uid_str)
            if uid_int // 100 != floor_num: continue      # id must belong to this floor
            ktype_val = wsbm.cell(row=row_ktype,  column=c_idx).value
            st_val    = wsbm.cell(row=row_status, column=c_idx).value
            ktype = str(ktype_val).strip() if ktype_val and str(ktype_val).strip() not in ("--", "None") else "--"
            raw_st = str(st_val).strip() if st_val and str(st_val).strip() not in ("--", "None") else "--"
            st = live_status.get(uid_int, STATUS_NORM.get(raw_st.upper(), raw_st))
            if str(st).upper() not in STATUS_MAP: st = "--"
            units.append({"uid": uid_str, "ktype": ktype, "status": st})

        if len(units) < 2: continue
        if floor_num in seen_floors:
            warn("Floor " + str(floor_num) + " appears more than once in the building matrix - later block ignored")
            continue
        seen_floors[floor_num] = True
        BUILDING_MATRIX.append({"floor": floor_num, "units": units})

    print("  -> " + str(len(BUILDING_MATRIX)) + " floors parsed")

    matrix_ids = set()
    for f in BUILDING_MATRIX:
        for u in f["units"]:
            matrix_ids.add(int(u["uid"]))
    orphans = sorted(set(owner) - matrix_ids)
    if orphans:
        warn(str(len(orphans)) + " unit(s) are in a container but not in the building matrix: " +
             ", ".join(str(o) for o in orphans[:12]) + ("..." if len(orphans) > 12 else ""))
    unassigned = sorted(int(u["uid"]) for f in BUILDING_MATRIX for u in f["units"]
                        if u["ktype"] != "--" and int(u["uid"]) not in owner)
    if unassigned:
        warn(str(len(unassigned)) + " kitchen(s) in the matrix are not in any container: " +
             ", ".join(str(o) for o in unassigned[:12]) + ("..." if len(unassigned) > 12 else ""))
else:
    warn("Building matrix sheet not found")

# ── Sanity guard ───────────────────────────────────────────────────
html = HTML_IN.read_text(encoding="utf-8")
prev = re.search(r"// DATA_START.*?// DATA_END", html, flags=re.DOTALL)
prev_ctn = len(re.findall(r"\bnum:", prev.group(0))) if prev else 0
prev_flr = len(re.findall(r"\bfloor:", prev.group(0))) if prev else 0

stop = []
if not CONTAINERS: stop.append("0 containers parsed")
if SH_BM and not BUILDING_MATRIX: stop.append("0 floors parsed from the building matrix")
if prev_ctn >= 5 and len(CONTAINERS) < prev_ctn * 0.7:
    stop.append("container count dropped from " + str(prev_ctn) + " to " + str(len(CONTAINERS)))
if prev_flr >= 5 and len(BUILDING_MATRIX) < prev_flr * 0.7:
    stop.append("floor count dropped from " + str(prev_flr) + " to " + str(len(BUILDING_MATRIX)))

log_lines = ["111W dashboard update - " + datetime.now().strftime("%Y-%m-%d %H:%M"),
             "Source: " + str(XL),
             "Containers: " + str(len(CONTAINERS)) + "  Floors: " + str(len(BUILDING_MATRIX)),
             ""]
log_lines += ["WARNING: " + w for w in WARNINGS] or ["No warnings."]

if stop and not FORCE:
    log_lines += ["", "PUBLISH BLOCKED:"] + ["  - " + s for s in stop]
    LOG_OUT.write_text("\n".join(log_lines), encoding="utf-8")
    print("")
    print("=" * 62)
    print("PUBLISH BLOCKED - the dashboard was NOT changed, committed or pushed.")
    for s in stop: print("  - " + s)
    print("")
    print("This usually means a column was renamed, a sheet was added or")
    print("reordered, or rows were inserted in the building matrix.")
    print("Check " + LOG_OUT.name + ", fix the spreadsheet, and run again.")
    print("If the drop is intentional, re-run with:  --force")
    print("=" * 62)
    if sys.platform == "win32": input("\nPress Enter to close...")
    sys.exit(1)

if not prev:
    sys.exit("ERROR: DATA markers not found in HTML. Expected: // DATA_START ... // DATA_END")

# ── Emit ───────────────────────────────────────────────────────────
_now = datetime.now()
now_str = _now.strftime("%b ") + str(_now.day) + _now.strftime(", %Y ") + str(_now.hour % 12 or 12) + _now.strftime(":%M ") + ("AM" if _now.hour < 12 else "PM")

TOTAL_KITCHENS = sum(1 for f in BUILDING_MATRIX for u in f["units"] if u["ktype"] != "--")

parts = []
parts.append("// DATA_START (auto-generated " + now_str + ")")
parts.append("const CONTAINERS = [")
for c in CONTAINERS:
    parts.append("  { num:" + json.dumps(c["num"]) + ", status:" + json.dumps(c["status"]) +
        ", week:" + json.dumps(c["week"]) + ", units:" + json.dumps(c["units"]) +
        ", unitQty:" + str(c["unitQty"]) + ", vessel:" + json.dumps(c["vessel"]) +
        ", loadDate:" + json.dumps(c["loadDate"]) + ", shipDate:" + json.dumps(c["shipDate"]) +
        ", portArrival:" + json.dumps(c["portArrival"]) + ", delivery:" + json.dumps(c["delivery"]) +
        ", deliveryISO:" + json.dumps(c["deliveryISO"]) +
        ", confirmed:" + ("true" if c["confirmed"] else "false") +
        ", floors:" + json.dumps(c["floors"]) + ", note:" + json.dumps(c["note"]) +
        ", code:" + json.dumps(c.get("code","—")) + ", delivTime:" + json.dumps(c.get("delivTime","—")) + " },")
parts.append("];")
parts.append("")
parts.append("const UNIT_STATUS = {")
for uid, st in UNIT_STATUS.items():
    parts.append("  " + str(uid) + ": " + json.dumps(st) + ",")
parts.append("};")
parts.append("")
parts.append("const TRANSIT_PERF = [")
for r in TRANSIT_PERF:
    parts.append("  { label:" + json.dumps(r["label"]) + ", value:" + json.dumps(r["value"]) + " },")
parts.append("];")
parts.append("")
parts.append("const LAST_UPDATED = " + json.dumps(now_str) + ";")
parts.append("const TOTAL_KITCHENS = " + str(TOTAL_KITCHENS) + ";")
parts.append("const DATA_WARNINGS = " + json.dumps(WARNINGS) + ";")
parts.append("")
parts.append("const BUILDING_MATRIX = [")
for f in BUILDING_MATRIX:
    uj = ", ".join("{uid:" + json.dumps(u["uid"]) + ",ktype:" + json.dumps(u["ktype"]) +
        ",status:" + json.dumps(u["status"]) + "}" for u in f["units"])
    parts.append("  { floor:" + str(f["floor"]) + ", units:[" + uj + "] },")
parts.append("];")
parts.append("// DATA_END")

data_block = "\n".join(parts)

html_out = re.sub(r"// DATA_START.*?// DATA_END", lambda m: data_block, html, flags=re.DOTALL)
HTML_OUT.write_text(html_out, encoding="utf-8")

if stop and FORCE:
    log_lines += ["", "GUARD OVERRIDDEN WITH --force:"] + ["  - " + s for s in stop]
LOG_OUT.write_text("\n".join(log_lines), encoding="utf-8")

print("Dashboard updated: " + HTML_OUT.name)
print("  " + str(len(CONTAINERS)) + " containers | " + str(len(UNIT_STATUS)) +
    " unit overrides | " + str(len(TRANSIT_PERF)) + " transit rows | " +
    str(len(BUILDING_MATRIX)) + " floors | " + str(TOTAL_KITCHENS) + " kitchens")
if WARNINGS:
    print("  " + str(len(WARNINGS)) + " warning(s) - see " + LOG_OUT.name)

import subprocess, os
os.chdir(HERE)
subprocess.run(["git", "add", "111W_Container_Dashboard.html", "config.js"], check=False)
subprocess.run(["git", "commit", "-m", "Dashboard update " + now_str], check=False)
result = subprocess.run(["git", "push"], check=False, capture_output=True, text=True)
if result.returncode == 0:
    print("Pushed to GitHub successfully.")
else:
    print("Git push failed: " + result.stderr.strip())

if sys.platform == "win32":
    input("\nPress Enter to close...")
